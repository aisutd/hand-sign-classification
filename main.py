import cv2
import mediapipe as mp
from openpyxl import Workbook
from os import listdir

# Settings

# The name of the data excel file
name = 'data'

# Subtracts each position from the first hand position
relative = False


# Dataset
# No data for J and Z
letters = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y']
datasets = {}
for letter in letters:
    d = letter + '/'
    datasets[letter] = ['data/A/' + d + x for x in listdir('data/A/' + d)] + \
                       ['data/B/' + d + x for x in listdir('data/B/' + d)] + \
                       ['data/C/' + d + x for x in listdir('data/C/' + d)] + \
                       ['data/D/' + d + x for x in listdir('data/D/' + d)] + \
                       ['data/E/' + d + x for x in listdir('data/E/' + d)]

# Excel
wb = Workbook()
del wb['Sheet']
# ws = wb.active
# Use sheet for each letter
def get_letter(x):
    if x > 25:
        return chr(ord('A') + x // 26 - 1) + chr(ord('A') + (x % 26))
    return chr(ord('A') + x)

# Training
mp_drawing = mp.solutions.drawing_utils
mp_hands = mp.solutions.hands
with mp_hands.Hands(
    static_image_mode=True,
    max_num_hands=1,
    min_detection_confidence=0.5) as hands:
    for letter in letters:
        print(letter + ' data')
        row = 0
        ws = wb.create_sheet(letter)
        data = datasets[letter]
        l = str(len(data))

        for file in data:
            image = cv2.flip(cv2.imread(file), 1)
            results = hands.process(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
            if not results.multi_hand_landmarks:
                continue

            row = row + 1
            print(str(row) + ' / ' + l)
            for hand_landmarks in results.multi_hand_landmarks:
                p0 = relative and (hand_landmarks.landmark[0]) or {'x':0,'y':0,'z':0}
                for i in range(0, len(hand_landmarks.landmark)):
                    p = hand_landmarks.landmark[i]
                    ws[get_letter(i * 3 + 0) + str(row)] = p.x - p0.x
                    ws[get_letter(i * 3 + 1) + str(row)] = p.y - p0.y
                    ws[get_letter(i * 3 + 2) + str(row)] = p.z - p0.z

        print('Finished ' + letter + ': ' + str(row) + ' / ' + l)

wb.save(name + ".xlsx")

# Real-time Classification
cap = cv2.VideoCapture(0)
with mp_hands.Hands(
        max_num_hands=1,
        min_detection_confidence=0.25,
        min_tracking_confidence=0.25) as hands:
    while cap.isOpened():
        success, image = cap.read()
        if not success:
            print("Ignoring empty camera frame.")
            # If loading a video, use 'break' instead of 'continue'.
            continue

        image = cv2.cvtColor(cv2.flip(image, 1), cv2.COLOR_BGR2RGB)
        image.flags.writeable = False
        cv2.imwrite("DetectionResults.jpg", image)
        cv2.imshow('MediaPipe Hands', image)

        results = hands.process(image)

        image.flags.writeable = True
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
        if results.multi_hand_landmarks:
            points = results.multi_hand_landmarks

            for hand_landmarks in results.multi_hand_landmarks:
                mp_drawing.draw_landmarks(
                    image, hand_landmarks, mp_hands.HAND_CONNECTIONS)

        cv2.imshow('MediaPipe Hands', image)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
cap.release()
cv2.destroyAllWindows()